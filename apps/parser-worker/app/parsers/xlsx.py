from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.schemas import ParsedElement


def parse_xlsx(path: Path, max_rows: int, max_elements: int) -> list[ParsedElement]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    elements: list[ParsedElement] = []
    total_rows = 0
    try:
        for sheet in workbook.worksheets:
            header: list[str] | None = None
            for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
                total_rows += 1
                if total_rows > max_rows:
                    raise ValueError("电子表格行数超过限制")
                normalized = [_stringify(value) for value in values]
                if not any(normalized):
                    continue
                if header is None:
                    header = normalized
                    element_type = "table_header"
                else:
                    element_type = "table_row"
                elements.append(
                    ParsedElement(
                        text="\t".join(normalized),
                        element_type=element_type,
                        sheet=sheet.title,
                        metadata={"row": row_number, "headers": header},
                    )
                )
                if len(elements) > max_elements:
                    raise ValueError("解析结果元素数量超过限制")
    finally:
        workbook.close()
    return elements


def _stringify(value: Any) -> str:
    return "" if value is None else str(value)
